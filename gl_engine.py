from __future__ import annotations
from typing import Dict, List, Optional, Tuple
import re
import io
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
HR_KEYWORDS = [
    "salary", "salaries", "payroll", "wages", "bonus", "gratuity", "leave", "provident", "pension",
    "eobi", "medical", "recruitment", "training", "staff", "employee", "overtime", "allowance",
    "compensation", "pay", "hr", "human resource"
]
MANUAL_KEYWORDS = [
    "manual", "adjust", "adjustment", "reclass", "reclassification", "provision", "accrual",
    "true up", "correction", "reversal", "topside", "sweep", "plug"
]

FIELD_SYNONYMS = {
    "date": ["gl date","posting date","entry date","transaction date","document date","voucher date","journal date","date","posted date","created date","effective date","batch date"],
    "debit": ["debit","dr","debit amount","debit amt","local debit","dr amount"],
    "credit": ["credit","cr","credit amount","credit amt","local credit","cr amount"],
    "amount": ["amount","net amount","journal amount","txn amount","value","local amount","signed amount"],
    "account": ["gl account","account","account code","account number","object account","natural account","ledger account","coa","account name","gl code","account desc","head","gl head","description","end account"],
    "party": ["party","vendor","customer","supplier","employee","subledger","business partner","bp","name","party name","vendor name","customer name","counterparty","third party","subledger name"],
    "user": ["user","posted by","created by","entered by","prepared by","maker","login","username","operator"],
    "voucher": ["voucher","voucher no","document number","document no","reference","ref","batch number","journal id","journal number","entry no","txn id","transaction id","document id","batch no"],
    "narration": ["narration","description","remarks","memo","comment","explanation","je explanation","text","line description","header text","item text","system descriptions"],
}

ALL_FIELDS = ["date","account","debit","credit","amount","party","user","voucher","narration"]
REQUIRED_FIELDS = ["date","account"]


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower()).strip()


def _non_null_ratio(series: pd.Series) -> float:
    return float(series.notna().mean()) if len(series) else 0.0


def _text_ratio(series: pd.Series) -> float:
    sample = series.dropna().astype(str).head(80)
    if len(sample) == 0:
        return 0.0
    return float(sample.str.contains(r"[A-Za-z]").mean())


def _numeric_ratio(series: pd.Series) -> float:
    coerced = pd.to_numeric(series.astype(str).str.replace(",", "", regex=False).str.replace("(", "-", regex=False).str.replace(")", "", regex=False), errors="coerce")
    return float(coerced.notna().mean()) if len(series) else 0.0


def _date_ratio(series: pd.Series) -> float:
    parsed = pd.to_datetime(series, errors="coerce")
    return float(parsed.notna().mean()) if len(series) else 0.0


def _unique_ratio(series: pd.Series) -> float:
    non_null = series.dropna()
    if len(non_null) == 0:
        return 0.0
    return min(1.0, float(non_null.nunique()) / max(1.0, len(non_null)))


def _header_score(col: str, field: str) -> float:
    c = _norm(col)
    score = 0.0
    for syn in FIELD_SYNONYMS[field]:
        s = _norm(syn)
        if c == s:
            score = max(score, 1.0)
        elif s in c:
            score = max(score, 0.8)
    if field == "date" and "date" in c:
        score = max(score, 0.75)
    if field == "account" and ("account" in c or "ledger" in c or "head" in c):
        score = max(score, 0.75)
    return score


def _value_score(series: pd.Series, field: str) -> float:
    if field == "date":
        return _date_ratio(series)
    if field in ["debit","credit","amount"]:
        return _numeric_ratio(series)
    if field == "account":
        return min(1.0, 0.45 * _text_ratio(series) + 0.55 * _unique_ratio(series))
    if field in ["party","user","voucher","narration"]:
        return min(1.0, 0.4 * _text_ratio(series) + 0.3 * _non_null_ratio(series) + 0.3 * _unique_ratio(series))
    return 0.0


def score_column(df: pd.DataFrame, col: str, field: str) -> float:
    return 0.65 * _header_score(col, field) + 0.35 * _value_score(df[col], field)


def infer_mapping(df: pd.DataFrame) -> Tuple[Dict[str, Optional[str]], Dict[str, float]]:
    mapping = {}
    confidence = {}
    used = set()
    for field in ALL_FIELDS:
        choices = []
        for col in df.columns:
            if col in used:
                continue
            choices.append((col, score_column(df, col, field)))
        choices.sort(key=lambda x: x[1], reverse=True)
        if choices:
            col, sc = choices[0]
            threshold = 0.45 if field in REQUIRED_FIELDS else 0.35
            if sc >= threshold:
                mapping[field] = col
                confidence[field] = round(sc, 2)
                used.add(col)
            else:
                mapping[field] = None
                confidence[field] = round(sc, 2)
        else:
            mapping[field] = None
            confidence[field] = 0.0
    return mapping, confidence


def _coerce_number(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace(",", "", regex=False).str.replace("(", "-", regex=False).str.replace(")", "", regex=False)
    return pd.to_numeric(s, errors="coerce")


def normalize_gl(df: pd.DataFrame, mapping_override: Optional[Dict[str, Optional[str]]] = None):
    work = df.copy()
    work["_source_row_num"] = np.arange(len(work)) + 2
    auto_mapping, confidence = infer_mapping(work)
    mapping = dict(auto_mapping)
    if mapping_override:
        for k, v in mapping_override.items():
            mapping[k] = None if v in ["", None, "__none__"] else v

    warnings = []
    if mapping["date"] is None:
        raise ValueError("Date column detect nahi hui. Custom Mapping se date select karo.")
    if mapping["account"] is None:
        raise ValueError("Account column detect nahi hui. Custom Mapping se GL account select karo.")

    work["gl_date"] = pd.to_datetime(work[mapping["date"]], errors="coerce")
    work["_debit"] = _coerce_number(work[mapping["debit"]]).fillna(0.0) if mapping.get("debit") else 0.0
    work["_credit"] = _coerce_number(work[mapping["credit"]]).fillna(0.0) if mapping.get("credit") else 0.0

    if mapping.get("amount"):
        amt = _coerce_number(work[mapping["amount"]]).fillna(0.0)
        if mapping.get("debit") is None and mapping.get("credit") is None:
            work["_debit"] = amt.clip(lower=0)
            work["_credit"] = (-amt.clip(upper=0)).abs()
            warnings.append("Debit/Credit columns nahi mili; Amount se split kiya gaya.")
        elif mapping.get("debit") is None:
            work["_debit"] = amt.clip(lower=0)
            warnings.append("Debit column nahi mili; Amount se infer ki gayi.")
        elif mapping.get("credit") is None:
            work["_credit"] = (-amt.clip(upper=0)).abs()
            warnings.append("Credit column nahi mili; Amount se infer ki gayi.")

    if float(pd.Series(work["_debit"]).abs().sum() + pd.Series(work["_credit"]).abs().sum()) == 0:
        raise ValueError("Amount columns detect nahi hui. Debit/Credit ya signed Amount map karo.")

    work["Amount"] = work["_debit"] - work["_credit"]
    work["account"] = work[mapping["account"]].fillna("Unknown").astype(str).str.strip()

    def safe_pick(field: str, fallback: str):
        col = mapping.get(field)
        if col and col in work.columns:
            return work[col].fillna(fallback).astype(str).str.strip()
        return pd.Series([fallback] * len(work), index=work.index)

    work["party"] = safe_pick("party", "Unknown")
    work["user"] = safe_pick("user", "Unknown")
    work["journal_id"] = safe_pick("voucher", "Unknown")
    work["narration"] = safe_pick("narration", "")
    return work, mapping, confidence, warnings


def _risk_label(score: int) -> str:
    if score >= 85:
        return "High"
    if score >= 70:
        return "Medium"
    return "Low"


def generate_samples(valid: pd.DataFrame, min_coverage: float = 0.15):
    df = valid.copy()
    if df.empty:
        return pd.DataFrame(), []

    sampled_parts = []
    summaries = []

    for account, grp in df.groupby("account", dropna=False):
        grp = grp.sort_values(["risk_score", "amount_abs"], ascending=[False, False]).copy()
        head_balance = float(grp["amount_abs"].sum())
        if head_balance <= 0:
            continue
        target = head_balance * min_coverage

        chosen = grp[grp["risk_label"] == "High"].copy()
        coverage = float(chosen["amount_abs"].sum()) if not chosen.empty else 0.0

        if coverage < target:
            remain = grp.loc[~grp.index.isin(chosen.index)]
            med = remain[remain["risk_label"] == "Medium"]
            for idx, row in med.iterrows():
                chosen = pd.concat([chosen, med.loc[[idx]]])
                coverage += float(row["amount_abs"])
                if coverage >= target:
                    break

        if coverage < target:
            remain = grp.loc[~grp.index.isin(chosen.index)].sort_values("amount_abs", ascending=False)
            for idx, row in remain.iterrows():
                chosen = pd.concat([chosen, remain.loc[[idx]]])
                coverage += float(row["amount_abs"])
                if coverage >= target:
                    break

        if chosen.empty:
            continue

        chosen["head_balance"] = head_balance
        chosen["coverage_pct_of_head"] = chosen["amount_abs"].cumsum() / head_balance
        sampled_parts.append(chosen)

        summaries.append({
            "account": str(account),
            "head_balance": round(head_balance, 2),
            "sample_count": int(len(chosen)),
            "sample_amount": round(float(chosen["amount_abs"].sum()), 2),
            "coverage_pct": round(float(chosen["amount_abs"].sum()) / head_balance * 100, 1),
        })

    if not sampled_parts:
        return pd.DataFrame(), []

    sample_df = pd.concat(sampled_parts).sort_values(["account", "risk_score", "amount_abs"], ascending=[True, False, False]).copy()
    preview_cols = ["_source_row_num","journal_id","gl_date","account","party","user","Amount","amount_abs","risk_score","risk_label","sample_bucket","reasons","narration","head_balance","coverage_pct_of_head"]
    sample_df = sample_df[preview_cols]
    sample_df = sample_df.rename(columns={"gl_date": "date", "amount_abs": "sample_amount_abs", "_source_row_num": "source_row_num"})
    sample_df["date"] = sample_df["date"].dt.strftime("%Y-%m-%d")
    sample_df["reasons"] = sample_df["reasons"].apply(lambda x: ", ".join(list(x)) if isinstance(x, list) else str(x))
    return sample_df.reset_index(drop=True), summaries


def _cv(values: pd.Series) -> float:
    arr = values.astype(float)
    mean = float(arr.mean()) if len(arr) else 0.0
    std = float(arr.std(ddof=0)) if len(arr) else 0.0
    if abs(mean) < 1e-9:
        return 0.0 if std == 0 else 9.99
    return std / abs(mean)


def _fmt_money(x: float) -> str:
    return f"{x:,.0f}"


def _period_range_strings(valid: pd.DataFrame) -> List[str]:
    if valid.empty:
        return []
    start = valid["gl_date"].min().to_period("M")
    end = valid["gl_date"].max().to_period("M")
    return [str(p) for p in pd.period_range(start=start, end=end, freq="M")]


def _build_complete_monthly(df: pd.DataFrame, group_field: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    periods = _period_range_strings(df)
    groups = sorted(df[group_field].astype(str).fillna("Unknown").unique())
    monthly = df.groupby([group_field, "period"]).agg(
        entries=("journal_id", "count"),
        signed_amount=("Amount", "sum"),
        gross_movement=("amount_abs", "sum"),
        unique_parties=("party", "nunique"),
        avg_ticket=("amount_abs", "mean"),
        weekend_entries=("is_weekend", "sum"),
        month_end_entries=("is_month_end", "sum"),
        manual_entries=("manual_signal", "sum"),
    )
    full_index = pd.MultiIndex.from_product([groups, periods], names=[group_field, "period"])
    monthly = monthly.reindex(full_index).reset_index()

    fill_zero = [
        "entries", "signed_amount", "gross_movement", "unique_parties", "avg_ticket",
        "weekend_entries", "month_end_entries", "manual_entries"
    ]
    for col in fill_zero:
        monthly[col] = monthly[col].fillna(0.0)
    monthly["entries"] = monthly["entries"].astype(int)
    monthly["unique_parties"] = monthly["unique_parties"].astype(int)
    monthly["weekend_entries"] = monthly["weekend_entries"].astype(int)
    monthly["month_end_entries"] = monthly["month_end_entries"].astype(int)
    monthly["manual_entries"] = monthly["manual_entries"].astype(int)

    monthly["prior_signed_amount"] = monthly.groupby(group_field)["signed_amount"].shift(1).fillna(0.0)
    monthly["mom_change_pct"] = np.where(
        monthly["prior_signed_amount"].abs() > 0,
        (monthly["signed_amount"] - monthly["prior_signed_amount"]) / monthly["prior_signed_amount"].abs(),
        np.nan,
    )
    return monthly


def _detect_hr_accounts(valid: pd.DataFrame) -> List[str]:
    if valid.empty:
        return []

    acct_name_hit = valid["account"].astype(str).str.lower().apply(lambda x: any(k in x for k in HR_KEYWORDS))
    narration_hit = valid["narration"].astype(str).str.lower().apply(lambda x: any(k in x for k in HR_KEYWORDS))
    party_hit = valid["party"].astype(str).str.lower().apply(lambda x: any(k in x for k in HR_KEYWORDS))

    stats = valid.assign(
        acct_name_hit=acct_name_hit,
        narration_hit=narration_hit,
        party_hit=party_hit,
    ).groupby("account").agg(
        entries=("journal_id", "count"),
        acct_name_hit=("acct_name_hit", "max"),
        narration_hit_rate=("narration_hit", "mean"),
        party_hit_rate=("party_hit", "mean"),
    )

    selected = stats[
        (stats["acct_name_hit"])
        | (stats["narration_hit_rate"] >= 0.50)
        | (stats["party_hit_rate"] >= 0.50)
    ].index.tolist()
    return [str(x) for x in selected]


def _dynamic_observations(valid: pd.DataFrame, month_counts: pd.Series, risk_distribution: List[Dict], party_summary: List[Dict], warnings: List[str]) -> List[str]:
    observations: List[str] = []
    if valid.empty:
        return observations

    top_month = month_counts.idxmax() if len(month_counts) else "N/A"
    top_month_count = int(month_counts.max()) if len(month_counts) else 0
    month_end_share = float(valid["is_month_end"].mean() * 100)
    weekend_share = float(valid["is_weekend"].mean() * 100)
    manual_share = float(valid["manual_signal"].mean() * 100)
    high_cnt = int((valid["risk_label"] == "High").sum())
    high_amt = float(valid.loc[valid["risk_label"] == "High", "amount_abs"].sum())
    total_amt = max(float(valid["amount_abs"].sum()), 1.0)
    q = valid["amount_abs"].quantile(0.95)
    large_cluster = valid[valid["amount_abs"] >= q]
    top_account = (
        valid.groupby("account")["amount_abs"].sum().sort_values(ascending=False).head(1)
    )
    if len(top_account):
        acc_name = str(top_account.index[0])
        acc_amt = float(top_account.iloc[0])
        observations.append(
            f"{acc_name} carried the highest gross movement at {_fmt_money(acc_amt)}, representing {acc_amt / total_amt * 100:.1f}% of the reviewed population. This head should be treated as a key analytical focus area because movement concentration is elevated relative to the rest of the ledger."
        )
    observations.append(
        f"Posting cadence is concentrated in {top_month} with {top_month_count:,} journals, while {month_end_share:.1f}% of entries were posted in the month-end window and {weekend_share:.1f}% fell on weekends. This timing profile is consistent with increased cut-off and manual adjustment risk around reporting close."
    )
    observations.append(
        f"High-risk journals comprise {high_cnt:,} entries with gross movement of {_fmt_money(high_amt)} ({high_amt / total_amt * 100:.1f}% of total absolute movement). The exceptions were driven by value outliers, posting timing and behavioral deviations rather than a single generic rule breach."
    )
    if len(large_cluster):
        lc_amt = float(large_cluster["amount_abs"].sum())
        observations.append(
            f"The top 5% value band contributed {_fmt_money(lc_amt)} of movement, indicating that a relatively small stratum of journals is driving a disproportionate share of exposure. A stratified testing approach will therefore be more efficient than purely random sampling."
        )
    if manual_share > 0:
        observations.append(
            f"Narration screening identified potential manual-adjustment language in {manual_share:.1f}% of journals. Entries containing reclass, accrual, adjustment or reversal indicators should be linked to supporting schedules and post-close approval evidence."
        )
    if party_summary:
        top_party = party_summary[0]
        observations.append(
            f"Counterparty concentration is led by {top_party['party']} with movement of {_fmt_money(float(top_party['amount']))}. Where this reflects recurring service or payroll-linked spend, assurance can be strengthened through expectation-based monthly analytics and contract-to-payment corroboration."
        )
    if warnings:
        observations.append("Auto-mapping inferred some fields from the uploaded ledger. Before finalizing testing, management should agree the mapped columns and the exact source-row extract used for sample selection.")
    return observations[:6]


def _build_assurance(valid: pd.DataFrame) -> Tuple[List[Dict], pd.DataFrame, List[Dict], pd.DataFrame]:
    if valid.empty:
        empty_df = pd.DataFrame()
        return [], empty_df, [], empty_df

    account_month = _build_complete_monthly(valid, "account")
    account_month = account_month.rename(columns={
        "gross_movement": "abs_movement",
        "prior_signed_amount": "prior_month_signed_amount",
    })

    party_concentration = valid.groupby(["account", "party"])["amount_abs"].sum().reset_index()
    total_by_account = valid.groupby("account")["amount_abs"].sum().rename("total").reset_index()
    party_concentration = party_concentration.merge(total_by_account, on="account", how="left")
    party_concentration["share_pct"] = np.where(party_concentration["total"] > 0, party_concentration["amount_abs"] / party_concentration["total"], 0.0)
    top_party_share = party_concentration.groupby("account")["share_pct"].max().rename("top_party_share")

    account_roll = valid.groupby("account").agg(
        entries=("journal_id", "count"),
        head_balance=("Amount", "sum"),
        gross_movement=("amount_abs", "sum"),
        avg_ticket=("amount_abs", "mean"),
        median_ticket=("amount_abs", "median"),
        unique_parties=("party", "nunique"),
        weekend_rate=("is_weekend", "mean"),
        month_end_rate=("is_month_end", "mean"),
        manual_rate=("manual_signal", "mean"),
        high_risk_count=("risk_label", lambda s: int((s == "High").sum())),
    )
    monthly_cv = account_month.groupby("account")["abs_movement"].apply(_cv).rename("monthly_cv")
    monthly_peak = account_month.groupby("account")["abs_movement"].max().rename("peak_month_movement")
    latest_mom = account_month.sort_values(["account", "period"]).groupby("account").tail(1).set_index("account")["mom_change_pct"].rename("latest_mom_change")
    roll = account_roll.join([monthly_cv, monthly_peak, latest_mom, top_party_share], how="left").fillna(0)

    assurance_rows: List[Dict] = []
    for account, row in roll.sort_values("gross_movement", ascending=False).head(20).iterrows():
        stable = row["monthly_cv"] <= 0.35
        concentrated = row["top_party_share"] >= 0.5
        manual = row["manual_rate"] >= 0.1 or row["month_end_rate"] >= 0.25
        payroll_like = any(k in _norm(account) for k in HR_KEYWORDS)
        if payroll_like:
            rec = "Monthly payroll bridge with headcount / payroll register reconciliation"
            why = "The head appears payroll-related, usually recurring in nature and therefore suitable for expectation-based analytics supported by HR master data and payroll registers."
            steps = "Build month-wise movement by account, compare to prior month, compare employee-level or payroll register totals, investigate exceptional bonus / leave / final settlement items, and agree the analytical bridge to approved payroll reports."
        elif stable and concentrated:
            rec = "Monthly vendor concentration and expectation testing"
            why = "Movement is recurring with a dominant counterparty, allowing an expectation to be developed from contract cadence, recurring invoices or service agreements."
            steps = "Prepare a 12-month movement trend, identify the dominant party, compare monthly charges to contracted rates / recurring invoices, and investigate unusual spikes or off-cycle postings."
        elif stable:
            rec = "Monthly trend and reasonableness analytics"
            why = "The head shows relatively stable run-rate behavior, so assurance can be taken through expectation setting, monthly variance analysis and corroboration to routine business drivers."
            steps = "Summarize monthly movement, compute month-on-month variance, define an expected range using historical average, investigate spikes above tolerance and corroborate with supporting schedules."
        elif manual:
            rec = "Focused journal-entry testing supported by close-cycle analytics"
            why = "Timing and manual-posting indicators suggest the balance is more exposed to management override and cut-off risk than to simple trend deviation."
            steps = "Filter month-end, weekend and manual-adjustment journals, stratify by value and preparer, inspect approvals and support, and separately assess whether close entries reverse in the following month."
        else:
            rec = "Stratified analytical review with targeted exception follow-up"
            why = "The head does not display a clean recurring pattern, but population-wide analytics can still provide coverage when combined with outlier testing and focused corroboration of abnormal movements."
            steps = "Develop monthly and counterparty-level summaries, isolate top-value and unusual journals, compare actual movement with historical pattern and investigate outliers beyond tolerance thresholds."

        assurance_rows.append({
            "account": str(account),
            "head_balance": round(float(row["head_balance"]), 2),
            "gross_movement": round(float(row["gross_movement"]), 2),
            "entries": int(row["entries"]),
            "monthly_cv": round(float(row["monthly_cv"]), 2),
            "top_party_share_pct": round(float(row["top_party_share"] * 100), 1),
            "high_risk_count": int(row["high_risk_count"]),
            "recommended_assurance": rec,
            "why_it_can_work": why,
            "procedure_to_apply": steps,
            "key_risk_focus": "Payroll / employee-related" if payroll_like else ("Manual close postings" if manual else ("Recurring spend pattern" if stable else "Volatile / outlier-driven movement")),
        })

    hr_accounts = _detect_hr_accounts(valid)
    hr_df = valid[valid["account"].astype(str).isin(hr_accounts)].copy() if hr_accounts else pd.DataFrame(columns=valid.columns)
    hr_rows: List[Dict] = []
    if not hr_df.empty:
        hr_month = _build_complete_monthly(hr_df, "account")
        hr_month = hr_month.rename(columns={
            "signed_amount": "net_amount",
            "gross_movement": "gross_movement",
            "prior_signed_amount": "prior_net_amount",
        })
        hr_month = hr_month[[
            "account", "period", "entries", "net_amount", "gross_movement", "unique_parties",
            "avg_ticket", "prior_net_amount", "mom_change_pct", "weekend_entries", "month_end_entries", "manual_entries"
        ]]

        hr_summary = hr_df.groupby("account").agg(
            net_movement=("Amount", "sum"),
            gross_movement=("amount_abs", "sum"),
            entries=("journal_id", "count"),
            unique_parties=("party", "nunique"),
            manual_rate=("manual_signal", "mean"),
        )
        for account, row in hr_summary.sort_values("gross_movement", ascending=False).iterrows():
            acc_norm = _norm(account)
            if any(k in acc_norm for k in ["salary", "payroll", "wages"]):
                proc = "Monthly payroll bridge"
                why = "Base payroll usually follows a recurring monthly run-rate and can be corroborated to payroll summaries, approved headcount and salary amendments."
                apply = "Compare month-wise GL payroll movement with payroll register totals; investigate new joiners, leavers, increments, overtime and unpaid leave movements."
            elif any(k in acc_norm for k in ["bonus", "gratuity", "leave", "provident", "eobi", "pension"]):
                proc = "Expectation testing linked to entitlement drivers"
                why = "These balances are often driven by formula-based entitlements or known event triggers, so expectation building is feasible."
                apply = "Build full month-wise movement for the whole account, tie unusual months to approved bonus runs / actuarial postings / settlements, and reconcile total movement to HR support schedules."
            else:
                proc = "Monthly HR expense trend with employee / vendor lens"
                why = "The head behaves like an HR-related cost pool and can be assessed through complete month-wise trend analytics, counterparty review and corroboration to HR / admin drivers."
                apply = "Build a complete monthly roll-forward for the full account population, investigate spikes and dormant-month reactivations, analyze recurring parties, and corroborate abnormal months to HR or procurement evidence."
            hr_rows.append({
                "account": str(account),
                "net_movement": round(float(row["net_movement"]), 2),
                "gross_movement": round(float(row["gross_movement"]), 2),
                "entries": int(row["entries"]),
                "unique_parties": int(row["unique_parties"]),
                "manual_signal_pct": round(float(row["manual_rate"] * 100), 1),
                "recommended_hr_procedure": proc,
                "why_applicable": why,
                "procedure_application": apply,
                "assurance_view": "Assurance can be taken where the complete month-wise movement of the selected head can be independently expected, reconciled or explained through payroll / HR source records and approved exception reports.",
            })
    else:
        hr_month = pd.DataFrame(columns=[
            "account", "period", "entries", "net_amount", "gross_movement", "unique_parties",
            "avg_ticket", "prior_net_amount", "mom_change_pct", "weekend_entries", "month_end_entries", "manual_entries"
        ])

    return assurance_rows, account_month, hr_rows, hr_month


def analyze_gl(df: pd.DataFrame, mapping_override: Optional[Dict[str, Optional[str]]] = None):
    work, mapping, confidence, warnings = normalize_gl(df, mapping_override)
    valid = work.dropna(subset=["gl_date"]).copy()
    if valid.empty:
        raise ValueError("Date parse nahi hui. Date format ya mapping check karo.")

    valid["period"] = valid["gl_date"].dt.to_period("M").astype(str)
    valid["month"] = valid["gl_date"].dt.strftime("%b")
    valid["day"] = valid["gl_date"].dt.day
    valid["weekday"] = valid["gl_date"].dt.day_name()
    valid["amount_abs"] = valid["Amount"].abs()
    valid["manual_signal"] = valid["narration"].astype(str).str.lower().apply(lambda x: any(k in x for k in MANUAL_KEYWORDS))

    month_counts = valid.groupby("month").size().reindex(MONTH_ORDER, fill_value=0)
    avg_monthly = max(float(month_counts.mean()), 1.0)
    month_spike = month_counts / avg_monthly

    party_month = valid.groupby(["party", "month"])["amount_abs"].sum().unstack(fill_value=0).reindex(columns=MONTH_ORDER, fill_value=0)
    dormant_parties = set()
    for party, row in party_month.iterrows():
        arr = row.values
        for i in range(2, len(arr)):
            if arr[i] > 0 and arr[i - 1] == 0 and arr[i - 2] == 0:
                dormant_parties.add(party)
                break

    valid["is_round"] = (valid["amount_abs"] > 0) & ((valid["amount_abs"] % 100).round(2) == 0)
    valid["is_weekend"] = valid["weekday"].isin(["Saturday", "Sunday"])
    valid["is_month_end"] = valid["day"] >= 28
    party_median = valid.groupby("party")["amount_abs"].transform("median").replace(0, np.nan)
    valid["party_spike"] = valid["amount_abs"] > (party_median.fillna(valid["amount_abs"].median()) * 2.5)
    acct_median = valid.groupby("account")["amount_abs"].transform("median").replace(0, np.nan)
    valid["account_spike"] = valid["amount_abs"] > (acct_median.fillna(valid["amount_abs"].median()) * 3.0)
    valid["rare_user_on_account"] = valid.groupby(["account", "user"])["journal_id"].transform("count") == 1

    score = np.full(len(valid), 28.0)
    score += np.where(valid["amount_abs"] >= valid["amount_abs"].quantile(0.90), 18, 0)
    score += np.where(valid["is_round"], 8, 0)
    score += np.where(valid["is_weekend"], 12, 0)
    score += np.where(valid["is_month_end"], 10, 0)
    score += np.where(valid["party"].isin(dormant_parties), 12, 0)
    score += np.where(valid["party_spike"], 10, 0)
    score += np.where(valid["account_spike"], 12, 0)
    score += np.where(valid["rare_user_on_account"], 8, 0)
    score += np.where(valid["manual_signal"], 10, 0)
    score += valid["month"].map(month_spike).fillna(1).clip(0, 3).mul(3).to_numpy()
    valid["risk_score"] = np.clip(score.round().astype(int), 0, 99)
    valid["risk_label"] = valid["risk_score"].apply(_risk_label)

    q90 = valid["amount_abs"].quantile(0.90)

    def reasons(row):
        items = []
        if row["amount_abs"] >= q90:
            items.append("Large value")
        if row["is_round"]:
            items.append("Round amount")
        if row["is_weekend"]:
            items.append("Weekend posting")
        if row["is_month_end"]:
            items.append("Month-end entry")
        if row["party"] in dormant_parties:
            items.append("Dormant party reactivation")
        if row["party_spike"]:
            items.append("Party movement spike")
        if row["account_spike"]:
            items.append("Account pattern spike")
        if row["rare_user_on_account"]:
            items.append("Rare preparer on account")
        if row["manual_signal"]:
            items.append("Manual adjustment wording")
        if not items:
            items.append("Behavior deviation")
        return items[:4]

    valid["reasons"] = valid.apply(reasons, axis=1)
    valid["sample_bucket"] = valid["risk_label"].map(lambda x: "100% High Risk" if x == "High" else ("Targeted Testing" if x == "Medium" else "Random Coverage"))

    monthly_entries = [{"month": m, "entries": int(month_counts[m])} for m in MONTH_ORDER if month_counts[m] > 0]
    movement_df = valid.groupby("month").agg(
        debit=("Amount", lambda s: float(s[s > 0].sum())),
        credit=("Amount", lambda s: float(abs(s[s < 0].sum())))
    ).reindex(MONTH_ORDER, fill_value=0.0)
    monthly_party_movement = [{"month": m, "debit": round(float(movement_df.loc[m, "debit"]), 2), "credit": round(float(movement_df.loc[m, "credit"]), 2)} for m in MONTH_ORDER if movement_df.loc[m].sum() > 0]

    risk_distribution = [
        {"name": "High", "value": int((valid["risk_label"] == "High").sum())},
        {"name": "Medium", "value": int((valid["risk_label"] == "Medium").sum())},
        {"name": "Low", "value": int((valid["risk_label"] == "Low").sum())},
    ]

    party_summary_raw = valid.groupby("party").agg(entries=("journal_id", "count"), amount=("amount_abs", "sum"))
    party_summary_raw["peak_month"] = valid.groupby(["party", "month"]).size().unstack(fill_value=0).reindex(columns=MONTH_ORDER, fill_value=0).idxmax(axis=1)
    med = party_summary_raw["amount"].median() or 1.0
    party_summary_raw["movement_pct"] = ((party_summary_raw["amount"] / med) - 1) * 100
    party_summary_raw["risk_label"] = np.where(party_summary_raw["movement_pct"] > 100, "High", np.where(party_summary_raw["movement_pct"] > 30, "Medium", "Low"))
    party_summary = []
    for party_name, row in party_summary_raw.sort_values("amount", ascending=False).head(10).iterrows():
        party_summary.append({
            "party": str(party_name),
            "entries": int(row["entries"]),
            "amount": round(float(row["amount"]), 2),
            "peak_month": str(row["peak_month"]),
            "movement_pct": round(float(row["movement_pct"]), 1),
            "risk_label": str(row["risk_label"]),
        })

    flagged_entries = []
    cols = ["_source_row_num", "journal_id","gl_date","account","party","user","Amount","risk_score","risk_label","reasons","sample_bucket","narration"]
    for _, row in valid.sort_values(["risk_score","amount_abs"], ascending=[False, False]).head(250)[cols].iterrows():
        flagged_entries.append({
            "source_row_num": int(row["_source_row_num"]),
            "journal_id": str(row["journal_id"]),
            "date": row["gl_date"].strftime("%Y-%m-%d") if pd.notna(row["gl_date"]) else "",
            "account": str(row["account"]),
            "party": str(row["party"]),
            "user": str(row["user"]),
            "amount": round(float(row["Amount"]), 2),
            "risk_score": int(row["risk_score"]),
            "risk_label": str(row["risk_label"]),
            "reasons": ", ".join(list(row["reasons"])),
            "sample_bucket": str(row["sample_bucket"]),
            "narration": str(row["narration"]),
        })

    sample_df, sample_summary = generate_samples(valid)
    sample_source_extract = df.iloc[(sample_df["source_row_num"] - 2).tolist()].copy() if not sample_df.empty else pd.DataFrame(columns=list(df.columns))
    if not sample_source_extract.empty:
        sample_source_extract.insert(0, "Source Row", sample_df["source_row_num"].tolist())
        sample_source_extract.insert(1, "Risk Label", sample_df["risk_label"].tolist())
        sample_source_extract.insert(2, "Risk Score", sample_df["risk_score"].tolist())
        sample_source_extract.insert(3, "Reasons", sample_df["reasons"].tolist())

    assurance_summary, assurance_monthly, hr_summary, hr_monthly = _build_assurance(valid)
    observations = _dynamic_observations(valid, month_counts, risk_distribution, party_summary, warnings)

    return {
        "summary": {
            "total_journals": int(len(valid)),
            "total_amount": round(float(valid["amount_abs"].sum()), 2),
            "high_risk_count": int((valid["risk_label"] == "High").sum()),
            "medium_risk_count": int((valid["risk_label"] == "Medium").sum()),
            "low_risk_count": int((valid["risk_label"] == "Low").sum()),
            "suggested_samples": int(len(sample_df)),
            "high_risk_samples": int((sample_df["risk_label"] == "High").sum()) if not sample_df.empty else 0,
            "medium_risk_samples": int((sample_df["risk_label"] == "Medium").sum()) if not sample_df.empty else 0,
            "low_risk_samples": int((sample_df["risk_label"] == "Low").sum()) if not sample_df.empty else 0,
        },
        "monthly_entries": monthly_entries,
        "monthly_party_movement": monthly_party_movement,
        "risk_distribution": risk_distribution,
        "top_observations": observations,
        "flagged_entries": flagged_entries,
        "party_summary": party_summary,
        "mapping": mapping,
        "mapping_confidence": confidence,
        "warnings": warnings,
        "source_columns": list(df.columns),
        "sample_records": sample_df.to_dict(orient="records"),
        "sample_summary": sample_summary,
        "sample_source_extract": sample_source_extract.to_dict(orient="records"),
        "assurance_summary": assurance_summary,
        "assurance_monthly": assurance_monthly.to_dict(orient="records"),
        "hr_summary": hr_summary,
        "hr_monthly": hr_monthly.to_dict(orient="records"),
    }


def _style_header(ws, row_idx: int, fill_color: str = "1F4E78"):
    header_fill = PatternFill("solid", fgColor=fill_color)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_widths(ws, widths: Dict[str, float]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def export_samples_to_excel(sample_records: List[Dict], sample_summary: List[Dict], sample_source_extract: Optional[List[Dict]] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Details"
    headers = ["Source Row","Journal ID","Date","Account","Party","User","Signed Amount","Absolute Amount","Risk Score","Risk Label","Sample Bucket","Reasons","Narration","Head Balance","Coverage % Running"]
    ws.append(headers)
    _style_header(ws, 1)
    for rec in sample_records:
        ws.append([
            rec.get("source_row_num"), rec.get("journal_id"), rec.get("date"), rec.get("account"), rec.get("party"), rec.get("user"),
            rec.get("Amount", rec.get("amount")), rec.get("sample_amount_abs"), rec.get("risk_score"), rec.get("risk_label"),
            rec.get("sample_bucket"), rec.get("reasons"), rec.get("narration"), rec.get("head_balance"), rec.get("coverage_pct_of_head")
        ])
    for col in ["G","H","N"]:
        for cell in ws[col][1:]:
            cell.number_format = '#,##0.00;[Red](#,##0.00)'
    for cell in ws["O"][1:]:
        cell.number_format = '0.0%'
    _apply_widths(ws, {"A":12,"B":16,"C":12,"D":26,"E":24,"F":18,"G":14,"H":14,"I":10,"J":12,"K":18,"L":36,"M":32,"N":16,"O":14})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Coverage Summary")
    ws2.append(["Account","Head Balance","Sample Count","Sample Amount","Coverage %"])
    _style_header(ws2, 1)
    for rec in sample_summary:
        ws2.append([rec.get("account"), rec.get("head_balance"), rec.get("sample_count"), rec.get("sample_amount"), rec.get("coverage_pct") / 100.0])
    for col in ["B","D"]:
        for cell in ws2[col][1:]:
            cell.number_format = '#,##0.00;[Red](#,##0.00)'
    for cell in ws2["E"][1:]:
        cell.number_format = '0.0%'
    _apply_widths(ws2, {"A":24,"B":16,"C":14,"D":16,"E":12})
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    ws3 = wb.create_sheet("Original GL Extract")
    source_rows = sample_source_extract or []
    if source_rows:
        headers3 = list(source_rows[0].keys())
        ws3.append(headers3)
        _style_header(ws3, 1, fill_color="7030A0")
        for rec in source_rows:
            ws3.append([rec.get(h) for h in headers3])
        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = ws3.dimensions
        for i, header in enumerate(headers3, start=1):
            col_letter = ws3.cell(1, i).column_letter
            width = max(12, min(28, len(str(header)) + 3))
            ws3.column_dimensions[col_letter].width = width
    else:
        ws3.append(["No sample source extract available"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def export_assurance_to_excel(assurance_summary: List[Dict], assurance_monthly: List[Dict], hr_summary: List[Dict], hr_monthly: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Assurance Summary"
    if assurance_summary:
        hdr = list(assurance_summary[0].keys())
        ws.append(hdr)
        _style_header(ws, 1, fill_color="0F766E")
        for rec in assurance_summary:
            ws.append([rec.get(h) for h in hdr])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ["B", "C"]:
            for cell in ws[col][1:]:
                cell.number_format = '#,##0.00;[Red](#,##0.00)'
        for col in ["E", "F"]:
            for cell in ws[col][1:]:
                cell.number_format = '0.0'
        _apply_widths(ws, {"A":28,"B":16,"C":16,"D":10,"E":12,"F":16,"G":14,"H":42,"I":48,"J":58,"K":24})
    else:
        ws.append(["No assurance summary available"])

    ws2 = wb.create_sheet("Monthly Analytics")
    if assurance_monthly:
        hdr2 = list(assurance_monthly[0].keys())
        ws2.append(hdr2)
        _style_header(ws2, 1, fill_color="1D4ED8")
        for rec in assurance_monthly:
            ws2.append([rec.get(h) for h in hdr2])
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
        _apply_widths(ws2, {"A":28,"B":12,"C":10,"D":16,"E":16,"F":16,"G":14,"H":16,"I":18,"J":16})
        for col in ["D","E","F"]:
            for cell in ws2[col][1:]:
                cell.number_format = '#,##0.00;[Red](#,##0.00)'
        for cell in ws2["J"][1:]:
            cell.number_format = '0.0%'
    else:
        ws2.append(["No monthly analytics available"])

    ws3 = wb.create_sheet("HR Procedures")
    if hr_summary:
        hdr3 = list(hr_summary[0].keys())
        ws3.append(hdr3)
        _style_header(ws3, 1, fill_color="B45309")
        for rec in hr_summary:
            ws3.append([rec.get(h) for h in hdr3])
        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = ws3.dimensions
        _apply_widths(ws3, {"A":28,"B":16,"C":10,"D":12,"E":16,"F":32,"G":48,"H":58,"I":42})
    else:
        ws3.append(["No HR-related heads detected from the uploaded GL."])

    ws4 = wb.create_sheet("HR Monthly Analytics")
    if hr_monthly:
        hdr4 = list(hr_monthly[0].keys())
        ws4.append(hdr4)
        _style_header(ws4, 1, fill_color="7C3AED")
        for rec in hr_monthly:
            ws4.append([rec.get(h) for h in hdr4])
        ws4.freeze_panes = "A2"
        ws4.auto_filter.ref = ws4.dimensions
        _apply_widths(ws4, {"A":28,"B":12,"C":10,"D":16,"E":16,"F":14,"G":14,"H":16,"I":14,"J":14,"K":16,"L":14})
        for col in ["D","E","G","H"]:
            for cell in ws4[col][1:]:
                cell.number_format = '#,##0.00;[Red](#,##0.00)'
        for cell in ws4["I"][1:]:
            cell.number_format = '0.0%'
    else:
        ws4.append(["No HR monthly analytics available"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
